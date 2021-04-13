from openpyxl import Workbook
from openpyxl import load_workbook
import os
import multiWayTree as MWT
import const


def typeof(variate) -> str:
    """
    判断变量类型
    """
    type = None
    if isinstance(variate, int):
        type = "int"
    elif isinstance(variate, str):
        type = "str"
    elif isinstance(variate, float):
        type = "float"
    elif isinstance(variate, list):
        type = "list"
    elif isinstance(variate, tuple):
        type = "tuple"
    elif isinstance(variate, dict):
        type = "dict"
    elif isinstance(variate, set):
        type = "set"
    return type


def getType(variate):
    """
    返回变量类型
    """
    arr = {"int": "整数", "float": "浮点", "str": "字符串", "list": "列表", "tuple": "元组", "dict": "字典", "set": "集合"}
    vartype = typeof(variate)
    if not (vartype in arr):
        return "未知类型"
    return arr[vartype]


def createNameStrs(prevName=None, midName=None, nums=[1]*1, rjust=0, afterName=None, startPos=[1]*1) -> list:
    """
    创建名字序列；
    params:前缀、连接符、结束位置列表、对齐、后缀、开始位置列表
    """
    resultStrs = []
    string = ''
    # process stop pos
    if isinstance(nums, list):
        numeration = True
    elif isinstance(nums, int):
        numeration = False
        nums = [nums]
    # process start pos
    if isinstance(startPos, int):
        startPos = [startPos]
    if len(startPos) < len(nums):
        affixArr = [startPos[len(startPos)-1]]*(len(nums)-len(startPos))
        startPos.extend(affixArr)

    if len(nums) > 0:
        for i in range(len(nums)):
            # startPos[i]可能大于nums[i]
            if startPos[i] < nums[i]:
                for j in range(startPos[i], nums[i]+1):
                    string = (prevName if prevName else '') + (str(i + 1) if numeration else '') + \
                             (midName if midName else '') + str(j).rjust(rjust, '0') + \
                             (afterName if afterName else '')
                    resultStrs.append(string)
    else:
        pass
    return resultStrs


def loadExcel(path=None):
    """
    加载excel
    """
    excel = None
    if path is not None:
        excel = load_workbook(path)
    return excel


def createExcels(wb: Workbook, path=None, excelNames=[]):
    """
    保存批量复制的excel；
    params:excel，路径，名称列表
    """
    # 处理excelNames传入string的情况
    if isinstance(excelNames, str):
        excelNames = [excelNames]

    if excelNames != []:
        for i in range(len(excelNames)):
            wb.save((path if path else '') + '\\' + excelNames[i] + '.xlsx')
    return


def createSheets(wb: Workbook, sheetNames=[], deletePrev=True):
    """
    创建sheet；可接收字符串及list参数，只创建未存在sheet
    params:excel，名称列表，清除旧有数据flag
    """
    existNames = wb.sheetnames
    if isinstance(sheetNames, list):
        pass
    elif isinstance(sheetNames, str):
        sheetNames = [sheetNames]
    if sheetNames != []:
        for i in range(len(sheetNames)):
            if sheetNames[i] not in existNames:
                wb.create_sheet(sheetNames[i])
    if deletePrev:
        for i in range(len(existNames)):
            if existNames[i] not in sheetNames:
                del wb[existNames[i]]
    return


def changeSheetName(wb: Workbook, sheetName: str, newName: str) -> bool:
    """
    改变sheet名称
    """
    if sheetName not in wb.sheetnames:
        return False
    ws = wb[sheetName]
    ws.title = newName
    return True


def insertDataToPos(wb: Workbook, row, col, value, sheetNames=None):
    """
    在指定的sheet内插入数据；
    params:excel，行，列，值，名称列表
    """
    if sheetNames is None:
        sheetNames = wb.sheetnames
    elif isinstance(sheetNames, str):
        sheetNames = [sheetNames]
    elif isinstance(sheetNames, list):
        pass
    else:
        return False
    sheetNames = [i for i in sheetNames if i in wb.sheetnames]

    if isinstance(value, str):
        value = [value]*len(sheetNames)
    elif isinstance(value, list):
        if len(value) < len(sheetNames):
            affixArr = [value[-1]]*(len(sheetNames)-len(value))
            value.extend(affixArr)
    else:
        return False
    for i in range(len(sheetNames)):
        sh = wb[sheetNames[i]]
        sh.cell(row, col, value[i])
    return True
# 测试用例
# insertDataToPos(wb, 2, 1, 'test')
# insertDataToPos(wb, 1, 2, ['test1', 'test2'])
# insertDataToPos(wb, 2, 2, ['test1', 'test2', 'test3', 'test4'])
# insertDataToPos(wb, 3, 3, 'test0', 'Sheet')
# insertDataToPos(wb, 4, 4, 'test0', ['Sheet', 'test3'])
# insertDataToPos(wb, 5, 5, 'test0', ['Shet', 'test3'])


def makeDir(path):
    """
    创建目录
    """
    # 移除字符串头尾的字符（默认为空格或换行符）
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("\\")
    if os.path.exists(path):
        # 如果目录存在则不创建，并提示目录已存在
        print(path + '目录已存在')
        return False
    else:
        # 如果不存在则创建目录
        os.makedirs(path)
        return True


def removeDir(path):
    """
    移除目录
    """
    for i in os.listdir(path):
        path_file = os.path.join(path, i)
        if os.path.isfile(path_file):
            os.remove(path_file)
        else:
            removeDir(path_file)
    # os.removedirs()会逐级删除到不能再删除，直接删了本来之后删除的父文件夹导致报FileNotFoundError，和函数的递归逻辑不相容
    # 如果去掉rmdir()可以实现删除文件而不删除文件夹
    os.rmdir(path)
    return


def createMWTElement(e: MWT.MultiWayTree, path: str, name: str, wb: Workbook) -> str:
    """
    按MWT结点创建叶子（非公开API）
    """
    result = path
    if e.type == const.TYPE_FOLDER:
        if e.name == const.STR_INPUT:
            result = path + '\\' + name
            makeDir(result)
        else:
            result = path + '\\' + e.name
            makeDir(result)
    elif e.type == const.TYPE_EXCEL:
        createExcels(wb, path, name)
    return result


def createMWTFolders(tree: MWT.MultiWayTree, path: str, name: str, wb: Workbook) -> str:
    """
    按目录和名词输出MWT给定的文件结构
    """
    tempPath = createMWTElement(tree, path, name, wb)
    for child_of_tree in tree.child:
        createMWTFolders(child_of_tree, tempPath, name, wb)
    return tempPath
